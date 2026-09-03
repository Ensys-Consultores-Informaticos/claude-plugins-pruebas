"""Compara un .xlsx generado contra uno exportado por Gesia, celda a celda.

Es la prueba que valida la apuesta de generar el fichero desde cero en vez de
rellenar el que exporta el auditor. Si los valores salen identicos, el formato
esta resuelto y deja de ser una opinion.

    python probar_fidelidad.py --generado nuestro.xlsx --referencia gesia.xlsx

Distingue dos clases de diferencia:
  · VALOR   rompe la importacion. Cualquiera es un fallo.
  · ESTILO  cosmetico (anchos, fuente). Se informa y no suspende.
"""

from __future__ import annotations

import argparse
import sys

from openpyxl import load_workbook

from lib_cuestionario import salida_utf8

MAX_LISTA = 15  # diferencias que se listan; el resto se cuenta


def normalizar(valor):
    """Deja comparables los valores que openpyxl no devuelve tal cual.

    openpyxl NO deshace el escape `_x000D_` con el que Excel guarda un retorno
    de carro dentro de una celda: lo devuelve como texto literal. Si un fichero
    se escribio con `\\r\\n` y el otro se leyo con el escape a la vista, las dos
    cadenas son la misma celda y hay que verlas iguales.
    """
    if isinstance(valor, str):
        return valor.replace("_x000D_", "\r").replace("\r\n", "\n").replace("\r", "\n")
    return valor


def comparar(gen_path: str, ref_path: str) -> int:
    gen = load_workbook(gen_path)
    ref = load_workbook(ref_path)

    fallos: list[str] = []
    avisos: list[str] = []

    if gen.sheetnames != ref.sheetnames:
        fallos.append(f"VALOR hojas: {gen.sheetnames} != {ref.sheetnames}")
        print("\n".join(fallos))
        return 1

    hg, hr = gen[gen.sheetnames[0]], ref[ref.sheetnames[0]]

    if (hg.max_row, hg.max_column) != (hr.max_row, hr.max_column):
        fallos.append(
            f"VALOR dimensiones: {hg.max_row}x{hg.max_column} != {hr.max_row}x{hr.max_column}"
        )

    filas = min(hg.max_row, hr.max_row)
    columnas = min(hg.max_column, hr.max_column)
    distintas = 0
    for i in range(1, filas + 1):
        for j in range(1, columnas + 1):
            cg, cr = hg.cell(row=i, column=j), hr.cell(row=i, column=j)
            vg, vr = normalizar(cg.value), normalizar(cr.value)
            if vg != vr:
                distintas += 1
                if len(fallos) < MAX_LISTA:
                    fallos.append(
                        f"VALOR {cg.coordinate}: generado={vg!r:.70} referencia={vr!r:.70}"
                    )
            elif type(cg.value) is not type(cr.value):
                distintas += 1
                if len(fallos) < MAX_LISTA:
                    fallos.append(
                        f"VALOR {cg.coordinate}: tipo {type(cg.value).__name__} "
                        f"!= {type(cr.value).__name__} (valor {vg!r:.40})"
                    )

    # Estilo: informativo.
    for i in range(1, columnas + 1):
        letra = hg.cell(row=1, column=i).column_letter
        ag = hg.column_dimensions[letra].width
        ar = hr.column_dimensions[letra].width
        if (ag or 0) != (ar or 0):
            avisos.append(f"ESTILO ancho {letra}: {ag} != {ar}")
    fg, fr = hg.cell(row=1, column=1).font, hr.cell(row=1, column=1).font
    if fg.bold != fr.bold:
        avisos.append(f"ESTILO cabecera negrita: {fg.bold} != {fr.bold}")
    fg2, fr2 = hg.cell(row=2, column=7).font, hr.cell(row=2, column=7).font
    if (fg2.name, fg2.size) != (fr2.name, fr2.size):
        avisos.append(f"ESTILO fuente: {fg2.name} {fg2.size} != {fr2.name} {fr2.size}")

    celdas = filas * columnas
    print(f"Comparadas {celdas} celdas ({filas} filas x {columnas} columnas).")
    if fallos:
        print(f"\n{distintas} diferencias de VALOR:")
        for f in fallos:
            print("  " + f)
        if distintas > len(fallos):
            print(f"  … y {distintas - len(fallos)} mas")
    else:
        print("Sin diferencias de valor: los dos ficheros son el mismo dato.")

    if avisos:
        print(f"\n{len(avisos)} diferencias de ESTILO (no afectan a la importacion):")
        for a in avisos[:MAX_LISTA]:
            print("  " + a)
        if len(avisos) > MAX_LISTA:
            print(f"  … y {len(avisos) - MAX_LISTA} mas")

    return 1 if distintas else 0


def main() -> int:
    salida_utf8()
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--generado", required=True)
    p.add_argument("--referencia", required=True)
    args = p.parse_args()
    return comparar(args.generado, args.referencia)


if __name__ == "__main__":
    sys.exit(main())
