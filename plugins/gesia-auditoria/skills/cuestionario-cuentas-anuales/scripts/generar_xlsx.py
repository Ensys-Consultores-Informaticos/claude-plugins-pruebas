"""Escribe el .xlsx con el formato de exportacion de Gesia, listo para importar.

Copia ONCE columnas literalmente del expediente y calcula UNA: `Respuesta`.
Nada mas. En particular `CodigoRealizado` NO se deduce: en un expediente real hay tres
preguntas contestadas con esa columna vacia, asi que no guarda relacion con
`Respuesta` y rellenarla seria inventar dato.

    python generar_xlsx.py --cuestionario trabajo/cuestionario.json \
        --respuestas trabajo/respuestas.json --salida "AG20-01_CLIENTE_2024.xlsx"

Sin `--respuestas` copia las que ya tenga el expediente. Es lo que usa la
comprobacion de fidelidad para regenerar un export existente y compararlo.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lib_cuestionario import (COLUMNAS, RESPUESTAS, es_cabecera,
                              leer_cuestionario, salida_utf8, valor_celda)

HOJA = "Guia"

# Anchos observados en el export de Gesia de AG)20/01. Son resultado de un
# autoajuste, asi que en otro cuestionario Gesia produciria otros. Es
# COSMETICO: no afecta a la importacion, que lee valores. Se replican porque
# sale gratis y porque el fichero se parece mas al que el auditor conoce.
ANCHOS = {
    "Prioritaria": 11.19921875,
    "CodigoArea": 12.9296875,
    "CodigoOrden": 14.33203125,
    "Punto": 6.73046875,
    "DesglosePunto": 16.33203125,
    "Descripcion": 255.59765625,
    "CodigoClaseReferencia": 25.06640625,
    "CodigoReferencia": 19.19921875,
    "CodigoRealizado": 18.19921875,
    "Respuesta": 11.59765625,
    "Comentario": 255.59765625,
}  # `CodigoGuia` no lleva ancho en el export: se deja al valor por defecto.

# El export de Gesia sale con Aptos Narrow 11, que es la fuente por defecto del
# libro, no un estilo de celda. openpyxl no sabe cambiar la del tema (probado:
# tocar el estilo 'Normal' no surte efecto al releer), asi que se pone celda a
# celda. openpyxl deduplica estilos, o sea que no engorda el fichero.
FUENTE = Font(name="Aptos Narrow", size=11)
FUENTE_CABECERA = Font(name="Aptos Narrow", size=11, bold=True)


def cargar_respuestas(ruta: str | None) -> dict[str, int]:
    """`{CodigoOrden: respuesta}`. Vacio si no se pasa fichero."""
    if not ruta:
        return {}
    datos = json.loads(Path(ruta).read_text(encoding="utf-8"))
    filas = datos["respuestas"] if isinstance(datos, dict) else datos
    return {str(r["orden"]): int(r["respuesta"]) for r in filas}


def main() -> int:
    salida_utf8()
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--cuestionario", required=True)
    p.add_argument("--respuestas", help="Si falta, se copian las del expediente")
    p.add_argument("--salida", required=True)
    args = p.parse_args()

    cuestionario = leer_cuestionario(args.cuestionario)
    filas = cuestionario["filas"]
    respuestas = cargar_respuestas(args.respuestas)

    sin_responder = []
    if respuestas:
        for fila in filas:
            if es_cabecera(fila):
                continue
            if str(fila["CodigoOrden"]) not in respuestas:
                sin_responder.append(str(fila["CodigoOrden"]))
    if sin_responder:
        print(f"ABORTA: faltan {len(sin_responder)} preguntas por responder.")
        print("Ordenes sin respuesta: " + ", ".join(sin_responder[:20])
              + (" …" if len(sin_responder) > 20 else ""))
        print("Toda pregunta lleva respuesta; si no se puede contestar, va un 4 (Pendiente).")
        return 2

    fuera_de_dominio = sorted({v for v in respuestas.values() if v not in RESPUESTAS})
    if fuera_de_dominio:
        print(f"ABORTA: respuestas fuera del dominio 1-4: {fuera_de_dominio}")
        return 2

    libro = Workbook()
    hoja = libro.active
    hoja.title = HOJA

    hoja.append(COLUMNAS)
    for celda in hoja[1]:
        celda.font = FUENTE_CABECERA

    for fila in filas:
        cabecera = es_cabecera(fila)
        valores = []
        for columna in COLUMNAS:
            if columna == "Respuesta" and respuestas and not cabecera:
                # Las cabeceras de seccion no se responden nunca: su Respuesta
                # queda vacia, igual que en el export de Gesia.
                valores.append(respuestas[str(fila["CodigoOrden"])])
            else:
                valores.append(valor_celda(columna, fila.get(columna)))
        hoja.append(valores)
        for celda in hoja[hoja.max_row]:
            celda.font = FUENTE

    for i, columna in enumerate(COLUMNAS, start=1):
        if columna in ANCHOS:
            hoja.column_dimensions[get_column_letter(i)].width = ANCHOS[columna]

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    libro.save(salida)

    preguntas = [f for f in filas if not es_cabecera(f)]
    print(f"Escrito: {salida}")
    print(f"{len(filas)} filas = {len(filas) - len(preguntas)} cabeceras + {len(preguntas)} preguntas")
    if respuestas:
        reparto = {c: sum(1 for v in respuestas.values() if v == c) for c in sorted(RESPUESTAS)}
        print("Reparto de respuestas: "
              + " · ".join(f"{c} {RESPUESTAS[c]}: {n}" for c, n in reparto.items()))
    else:
        print("Sin --respuestas: se han copiado las del expediente.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
