"""Escribe el papel de trabajo de continuidad de saldos en Excel.

    python generar_papel.py --resultado resultado.json \
        --fecha-cierre 31/12/25 --salida "<expediente>/InformesGesia/..."

El formato reproduce el papel que el auditor hacia a mano (el AG)02-02 del
ejercicio anterior): cabecera con cliente, ejercicio, titulo y fuente; el bloque de
REF. P.T. / Realizado / Verificado a la derecha; y una fila por cuenta con
apertura, cierre y diferencia. Sin fila de totales.

Nada aqui lee el reloj: la fecha de cierre se pasa como argumento.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lib_continuidad import salida_utf8

FORMATO_EURO = "#,##0.00"

# Ancho de columna en caracteres. Medido sobre el papel del ejercicio anterior:
# la columna del nombre necesita sitio porque a nivel 3 digitos son epigrafes y
# a 9 son nombres de tercero, mas largos.
ANCHOS = {"A": 14, "B": 46, "C": 10, "D": 16, "E": 16, "F": 18}

# Relleno de las filas con hallazgo. No lo llevaba el papel manual, pero con 438
# filas y todas las diferencias a cero salvo dos, sin marca no se encuentran.
RELLENO = PatternFill("solid", fgColor="FFF2CC")


def _euro(v) -> str:
    """Importe con separador de miles, para los textos de aviso."""
    return format(float(v), ",.2f").replace(",", " ")


def main() -> int:
    salida_utf8()
    p = argparse.ArgumentParser()
    p.add_argument("--resultado", required=True)
    p.add_argument("--salida", required=True)
    p.add_argument("--fecha-cierre", required=True,
                   help="como se escribe en el papel, p.ej. 31/12/25")
    p.add_argument("--ref", default="AG)02",
                   help="referencia del papel en Gesia")
    p.add_argument("--fuente", default="Contabilidad de la entidad")
    args = p.parse_args()

    datos = json.loads(Path(args.resultado).read_text(encoding="utf-8"))
    meta = datos["meta"]
    filas = datos["filas"]
    con_hallazgo = {h["cuenta"] for h in datos["hallazgos"]}

    wb = Workbook()
    ws = wb.active
    ws.title = args.ref

    negrita = Font(bold=True)

    ws["A1"] = meta.get("cliente", "")
    ws["A1"].font = negrita
    ws["A2"] = "AUDITORIA A " + args.fecha_cierre
    ws["A3"] = "SALDOS DE APERTURA"
    ws["A4"] = "Fuente: " + args.fuente

    # El aviso de cuadre va ARRIBA, no en la segunda hoja: si la apertura de
    # balance no suma cero hay un importe que la comparacion no esta viendo, y el
    # papel no puede firmarse sin saberlo. Es lo primero que se lee al abrirlo.
    cuadre = datos.get("cuadre", {})
    if abs(cuadre.get("descuadre_apertura_balance", 0)) >= 0.005:
        aviso = (
            "AVISO: la apertura de balance no cuadra a cero. Descuadre "
            + _euro(cuadre["descuadre_apertura_balance"]) + " EUR: "
            + _euro(cuadre["por_cuentas_de_resultados"])
            + " en cuentas de resultados y "
            + _euro(cuadre["por_cuentas_no_reconocidas"])
            + " en cuentas que el expediente no reconoce. Detalle en la hoja "
              "Incidencias."
        )
        if abs(cuadre.get("sin_explicar", 0)) >= 0.005:
            aviso += (" QUEDAN " + _euro(cuadre["sin_explicar"])
                      + " SIN EXPLICAR: no firmes este papel.")
        c = ws["A5"]
        c.value = aviso
        c.font = Font(bold=True)
        c.fill = RELLENO

    ws["G1"], ws["H1"] = "REF. P.T.", args.ref
    ws["G2"], ws["H2"] = "Realizado", ""
    ws["G3"], ws["H3"] = "Verificado", ""
    for celda in ("G1", "G2", "G3"):
        ws[celda].font = negrita

    cab = [
        "CUENTA",
        "NOMBRE",
        "REF",
        meta["ejercicio"],
        meta["ejercicio_anterior"],
        "Diferencia Apertura " + meta["ejercicio"] + "-Cierre "
        + meta["ejercicio_anterior"],
    ]
    for j, texto in enumerate(cab, start=1):
        c = ws.cell(6, j, texto)
        c.font = negrita
        c.alignment = Alignment(wrap_text=True, vertical="bottom")

    for i, f in enumerate(filas, start=7):
        ws.cell(i, 1, f["cuenta"])
        ws.cell(i, 2, f["nombre"])
        ws.cell(i, 3, "")                      # REF, en blanco a proposito
        # None cuando la cuenta no existe en ese lado: celda VACIA, no cero. Un
        # cero dice "esta y vale cero"; el hueco dice "no esta", que es el
        # hallazgo. Confundirlos es el error mas facil de cometer aqui.
        for col, clave in ((4, "apertura"), (5, "cierre"), (6, "diferencia")):
            valor = f[clave]
            c = ws.cell(i, col)
            if valor is not None:
                c.value = valor
                c.number_format = FORMATO_EURO
        if f["cuenta"] in con_hallazgo:
            for col in range(1, 7):
                ws.cell(i, col).fill = RELLENO

    for col, ancho in ANCHOS.items():
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A7"

    _hoja_incidencias(wb, datos)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(salida)
    except PermissionError:
        # Pasa siempre que se regenera el papel con el fichero abierto en Excel,
        # que es lo normal cuando el auditor esta revisandolo. Sin este mensaje
        # sale un traceback de zipfile que no dice que hay que cerrar el fichero.
        print("ERROR: no se puede escribir " + str(salida))
        print("       El fichero esta abierto en Excel. Cierralo y repite.")
        return 2

    print("Escrito: " + str(salida))
    print("  " + str(len(filas)) + " cuentas, " + str(len(datos["hallazgos"]))
          + " con hallazgo")
    if wb.sheetnames[-1] == "Incidencias":
        print("  hoja Incidencias con lo que no cabe en el papel")
    return 0


def _hoja_incidencias(wb: Workbook, datos: dict) -> None:
    """Segunda hoja, solo si hay algo que contar.

    Lo que va aqui no son diferencias de importe, asi que no cabe en la hoja
    principal: cuentas del diario que no encajan en el expediente y cuentas de
    resultados con saldo de apertura. Si no hay nada, no se crea la hoja: una
    hoja vacia en un papel de trabajo invita a pensar que se ha mirado algo que
    no se ha mirado.
    """
    sin_mapeo = datos.get("sin_mapeo", [])
    resultados = datos.get("resultados_en_apertura", [])
    if not sin_mapeo and not resultados:
        return

    ws = wb.create_sheet("Incidencias")
    negrita = Font(bold=True)
    fila = 1

    if sin_mapeo:
        ws.cell(fila, 1, "Cuentas abiertas que no encajan en ninguna cuenta del "
                         "expediente").font = negrita
        fila += 1
        ws.cell(fila, 1, "CUENTA").font = negrita
        ws.cell(fila, 2, "SALDO APERTURA").font = negrita
        fila += 1
        for s in sin_mapeo:
            ws.cell(fila, 1, s["cuenta"])
            c = ws.cell(fila, 2, s["saldo"])
            c.number_format = FORMATO_EURO
            fila += 1
        fila += 1

    if resultados:
        ws.cell(fila, 1, "Cuentas de los grupos 6 y 7 con saldo de apertura "
                         "(fuera de la prueba)").font = negrita
        fila += 1
        ws.cell(fila, 1, "CUENTA").font = negrita
        ws.cell(fila, 2, "SALDO APERTURA").font = negrita
        fila += 1
        for r in resultados:
            ws.cell(fila, 1, r["cuenta"])
            c = ws.cell(fila, 2, r["saldo"])
            c.number_format = FORMATO_EURO
            fila += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18


if __name__ == "__main__":
    sys.exit(main())
