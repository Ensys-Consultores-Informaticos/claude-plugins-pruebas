"""Escribe el papel de trabajo de la prueba MUM en Excel.

    python generar_papel.py --muestra muestra.json --parametros parametros.json \
        --facturas facturas.json [--evaluacion evaluacion.json] \
        --salida "<expediente>/InformesGesia/FspMum/MUM <PRUEBA> <CLIENTE> <EJERCICIO>.xlsx" \
        --generado 2026-09-03

Una sola hoja, «Analisis muestra»: una fila por elemento seleccionado con sus
columnas de poblacion, el documento localizado y lo leido en el, y las tres
columnas de la MUM propuestas -saldo segun auditoria, error y tasa- mas la
observacion para ForSampling.

Un solo color: amarillo en la observacion cuando hay error o cuando el elemento
no se ha podido medir. Lo demas, sin color.

**No hay fila de totales, y es a proposito.** Sumar los errores de una MUM es
proyectar a ojo, y sumarlos con su signo esconde las incorrecciones al
cancelarse. La proyeccion la hace ForSampling con la muestra evaluada. Lo que se
imprime al ejecutar son los errores por exceso y por defecto POR SEPARADO.

No usa formulas: todo se calcula aqui y se escribe como valor. Y no lee el reloj:
la fecha de generacion entra por --generado, para que el papel se pueda
regenerar identico dentro de dos años.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib_fsp import (  # noqa: E402
    TOL_IMPORTE,
    VENTANA_DIAS,
    cargar_evaluacion,
    cargar_facturas,
    cargar_muestra,
    cargar_parametros,
    cruzar,
    columnas_de_muestra,
    parse_fecha,
    parse_importe,
    salida_utf8,
)
from lib_mum import (  # noqa: E402
    NOMBRE_TERMINO,
    comparar_con_auditor_mum,
    evaluar_mum,
    observacion_mum,
    recuento,
    termino_mayoritario,
)

FONT = "Arial"
FORMATO_EURO = '#,##0.00 "€";-#,##0.00 "€";"-"'
FORMATO_PCT = '0.00 "%";-0.00 "%";"-"'
AMARILLO = PatternFill("solid", fgColor="FFFF00")
CABECERA_FILL = PatternFill("solid", fgColor="1F4E78")
CABECERA_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
_BORDE_LADO = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_BORDE_LADO, right=_BORDE_LADO, top=_BORDE_LADO, bottom=_BORDE_LADO)


def _cabecera(ws, fila: int, textos: list[str]) -> None:
    for j, t in enumerate(textos, start=1):
        c = ws.cell(row=fila, column=j, value=t)
        c.font = CABECERA_FONT
        c.fill = CABECERA_FILL
        c.border = BORDE
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _anchos(ws, anchos: dict[int, float]) -> None:
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _eur(v: float | None) -> str:
    if v is None:
        return "—"
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _hoja(wb: Workbook, evaluados: list[dict], cols: dict, params: dict, generado: str) -> None:
    """La unica hoja: un elemento por fila, sus tres columnas de MUM y la observacion.

    Las tres columnas SI se rellenan, al contrario que los atributos de la prueba
    de cumplimiento, y la diferencia no es un descuido. Un atributo es un
    veredicto -Ok o no Ok- y rellenarlo es concluir, que no es del skill. Un
    importe segun el documento es una medida: decir que la factura pone 16.962,50
    es describir. Van rotuladas «(propuesto)» para que se vea de quien es la
    propuesta, y el auditor las adopta o las cambia.
    """
    ws = wb.active
    ws.title = "Análisis muestra"
    if not evaluados:
        ws["A1"] = "La muestra no trae elementos."
        return

    pr = params.get("parametros") or {}
    et = parse_importe(pr.get("ErrorTolerableValor"))

    def _dia(v):
        d = parse_fecha(v)
        return d.strftime("%d/%m/%Y") if d else (str(v) if v else "—")

    col_pob = [c for c in evaluados[0]["fila"].keys() if c.lower() not in ("seleccionado",)]
    col_doc = ["Fichero", "Nº factura (doc.)", "Fecha (doc.)", "Base (doc.)", "IVA (doc.)",
               "IRPF (doc.)", "Total (doc.)", "Casa por", "Días doc.–libros"]
    col_mum = ["Saldo auditoría (propuesto)", "Error (propuesto)", "% error (propuesto)",
               "Comparado con"]
    cab = col_pob + col_doc + col_mum + ["Observación propuesta"]

    ws["A1"] = f"Análisis de la muestra — {params.get('Prueba') or 'prueba MUM'}"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws["A2"] = (f"Prueba {params.get('MuestraId')} · MUM · área {params.get('Area')} · "
                f"ref. {params.get('Referencia')} · ejercicio "
                f"{_dia(params.get('FechaInicioAuditoria'))} — {_dia(params.get('FechaFinAuditoria'))} · "
                f"unidad de muestreo {pr.get('UnidadMuestreo') or '—'} · población "
                f"{pr.get('PoblacionNumElementos') or '—'} elementos · error tolerable de la prueba "
                f"{_eur(et)} · generado {generado}")
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A3"] = ("Propuesta del asistente: localiza el documento de cada elemento y mide lo que sostiene, "
                "en el término con el que contabiliza esta población. El auditor adopta o cambia cada importe.")
    ws["A3"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws["A4"] = ("Amarillo: hay diferencia con los libros, o el elemento no se ha podido medir. Este papel NO "
                "proyecta el error a la población, no lo compara con el error tolerable y no suma los errores "
                "entre sí: eso lo hace ForSampling, y los errores por exceso y por defecto no se cancelan.")
    ws["A4"].font = Font(name=FONT, italic=True, size=9, color="595959")

    _cabecera(ws, 6, cab)
    fila = 7
    for e in evaluados:
        f, fac, c = e["fila"], e["factura"], e["criterios"]
        valores = []
        for k in col_pob:
            v = f.get(k)
            if k == cols.get("importe"):
                v = parse_importe(v)
            elif k == cols.get("fecha"):
                d = parse_fecha(v)
                v = d.strftime("%d/%m/%Y") if d else v
            valores.append(v)
        if fac:
            casa = ", ".join(x for x, b in (("número", c["numero"]), ("tercero", c.get("tercero")),
                                            (f"importe ({c['importe']})", c["importe"]),
                                            ("fecha", c["fecha"])) if b)
            valores += [fac.get("fichero"), fac.get("numero"), fac.get("fecha"),
                        parse_importe(fac.get("base")), parse_importe(fac.get("iva")),
                        parse_importe(fac.get("irpf")), parse_importe(fac.get("total")),
                        casa, c["dias"]]
        else:
            valores += ["— no localizada", None, None, None, None, None, None, "", None]
        valores += [e["saldo_auditoria"], e["error"], e["tasa"],
                    NOMBRE_TERMINO.get(e["termino"], "—")]
        valores.append(e["observacion"])
        for j, v in enumerate(valores, start=1):
            cell = ws.cell(row=fila, column=j, value=v)
            cell.font = Font(name=FONT, size=9)
            cell.border = BORDE
            cell.alignment = Alignment(vertical="top", wrap_text=j == len(cab))
            if isinstance(v, float):
                cell.number_format = FORMATO_EURO
        # la tasa es un porcentaje, no euros
        ws.cell(row=fila, column=len(col_pob) + len(col_doc) + 3).number_format = FORMATO_PCT
        if e["error"] is None or abs(e["error"]) > TOL_IMPORTE:
            ws.cell(row=fila, column=len(cab)).fill = AMARILLO
        fila += 1

    ws.auto_filter.ref = f"A6:{get_column_letter(len(cab))}{fila - 1}"
    ws.freeze_panes = "A7"
    anchos = {i + 1: 14 for i in range(len(cab))}
    for i, k in enumerate(col_pob, start=1):
        if k in (cols.get("tercero"), "Acreedor"):
            anchos[i] = 30
    anchos[len(col_pob) + 1] = 34
    for k in range(4):
        anchos[len(col_pob) + len(col_doc) + k + 1] = 20
    anchos[len(cab)] = 95
    _anchos(ws, anchos)


def main() -> int:
    salida_utf8()
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--muestra", required=True)
    p.add_argument("--parametros", required=True)
    p.add_argument("--facturas", required=True)
    p.add_argument("--evaluacion", help="exportar_consulta(entidad='evaluacion', id=N), para comparar")
    p.add_argument("--salida", required=True)
    p.add_argument("--generado", required=True, help="fecha de generación, AAAA-MM-DD. Nada lee el reloj")
    args = p.parse_args()

    muestra = cargar_muestra(args.muestra)
    params = cargar_parametros(args.parametros)
    facturas = cargar_facturas(args.facturas)
    evaluacion = cargar_evaluacion(args.evaluacion)

    cols, _, _ = columnas_de_muestra(muestra)
    cruce = cruzar(muestra, facturas, cols)
    evaluados = evaluar_mum(cruce, cols)
    for e in evaluados:
        e["observacion"] = observacion_mum(e)
    comp = comparar_con_auditor_mum(evaluados, (evaluacion or {}).get("filas") or [], cols)
    r = recuento(evaluados)

    wb = Workbook()
    _hoja(wb, evaluados, cols, params, args.generado)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(salida)
    except PermissionError:
        print(f"ERROR: no se puede escribir {salida}. Si está abierto en Excel, ciérralo y repite.")
        return 2

    term = termino_mayoritario(evaluados)
    print(f"Papel escrito: {salida}")
    print(f"  elementos {r['elementos']} · unidades de muestreo con repeticiones {r['repeticiones']:g}"
          f" · con documento {r['con_documento']} · sin documento {r['sin_documento']}"
          f" · medidos {r['medidos']} · sin medir {r['sin_medir']}")
    print(f"  esta población contabiliza por: {NOMBRE_TERMINO.get(term, 'sin criterio claro (lo decide el auditor)')}")
    print(f"  sin diferencia {r['sin_error']} · con diferencia {r['con_error']}")
    if r["con_error"]:
        print(f"  incorrecciones POR SEPARADO, sin netear: {r['n_exceso']} por exceso suman "
              f"{_eur(r['suma_exceso'])} · {r['n_defecto']} por defecto suman {_eur(r['suma_defecto'])}")
        print("  NO son la proyección ni el error neto: la proyección la hace ForSampling con la muestra evaluada")
    for e in evaluados:
        if e["error"] is None or abs(e["error"]) > TOL_IMPORTE:
            ident = e["fila"].get(cols.get("id"), "?")
            print(f"  · elemento {ident}: {e['observacion']}")
    for par in cruce.get("candidatos_sueltos") or []:
        print(f"  · POSIBLE DIFERENCIA, no extravío: el elemento {par['id']} quedó sin documento y "
              f"{par['fichero']} sin elemento"
              + (", mismo tercero" if par["tercero"] else "")
              + (f", {par['dias']} día(s) de diferencia" if par["dias"] is not None else "")
              + (f". En libros {par['importe_libros']}, en el documento base {par['base']} / total "
                 f"{par['total']}: diferencia mínima {par['diferencia_minima']}"
                 if par["diferencia_minima"] is not None else "")
              + ". Revísalo a mano: si son el mismo, hay una diferencia real que declarar con "
                "poblacion_id en facturas.json")
    for fac in cruce["facturas_sin_fila"]:
        print(f"  · documento que no es de ningún elemento de la muestra: {fac.get('fichero')}"
              f" ({fac.get('proveedor') or '—'}, {fac.get('total') or '—'})")
    if comp:
        c = comp["recuento"]
        print(f"  frente al auditor: coinciden {c['coinciden']} · skill señala error/auditor 0 "
              f"{c['skill_error_auditor_cero']} · SKILL 0/AUDITOR ERROR {c['skill_cero_auditor_error']}"
              f" · los dos con error distinto {c['discrepan']} · sin comparar {c['sin_evaluar']}")
        for f in comp["filas"]:
            if f["estado"].startswith("EL SKILL DA 0") or f["estado"].startswith("los dos"):
                print(f"    · elemento {f['id']}: skill {_eur(f['error_skill'])} / auditor "
                      f"{_eur(f['error_auditor'])} — {f['observacion_auditor'] or ''}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
