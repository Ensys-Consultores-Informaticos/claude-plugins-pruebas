"""Escribe el papel de trabajo de la prueba de cumplimiento en Excel.

    python generar_papel.py --muestra muestra.json --parametros parametros.json \
        --facturas facturas.json --roles roles.json [--evaluacion evaluacion.json] \
        --salida "<expediente>/InformesGesia/FspCumplimiento/Cumplimiento <PRUEBA> <CLIENTE> <EJERCICIO>.xlsx" \
        --generado 2026-09-02

Una sola hoja, «Analisis muestra»: una fila por elemento seleccionado con sus
columnas de poblacion, el documento localizado y lo leido en el (numero, fecha,
base, IVA, retencion, total), como caso, la diferencia de importe, una columna
por atributo EN BLANCO y la observacion propuesta para ForSampling.

Las columnas de atributos van vacias a proposito: el skill no marca Ok ni pone
'auditor', las deja para quien firma y cuenta lo que ha visto en la observacion.

Un solo color: amarillo en la observacion cuando señala algo. Lo demas, sin color.

Lo que no cabe en la hoja se imprime al ejecutar: los recuentos, los elementos
con hallazgo, los documentos que no son de ningun elemento y, si se paso
--evaluacion, la comparacion con lo que puso el auditor. Leelo antes de entregar.

No usa formulas: todo se calcula aqui y se escribe como valor, para que el papel
se lea igual en Cowork y en la maquina del auditor. Y no lee el reloj: la fecha
de generacion entra por --generado, para que el papel se pueda regenerar identico.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib_fsp import (  # noqa: E402
    ROL_AUDITOR,
    TOL_IMPORTE,
    TOL_IVA,
    VENTANA_DIAS,
    cargar_evaluacion,
    cargar_facturas,
    cargar_json,
    cargar_muestra,
    cargar_parametros,
    comparar_con_auditor,
    cruzar,
    columnas_de_muestra,
    evaluar,
    parse_fecha,
    parse_importe,
    proponer_roles,
    salida_utf8,
)

FONT = "Arial"
FORMATO_EURO = '#,##0.00 "€";-#,##0.00 "€";"-"'
AMARILLO = PatternFill("solid", fgColor="FFFF00")
CABECERA_FILL = PatternFill("solid", fgColor="1F4E78")
CABECERA_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
_BORDE_LADO = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_BORDE_LADO, right=_BORDE_LADO, top=_BORDE_LADO, bottom=_BORDE_LADO)

NOMBRES_ROL = {
    "documento": "el skill comprueba que el documento existe y es el del apunte",
    "calculo": ("el skill comprueba la aritmética del documento: base + IVA = total, "
                "y con retención de IRPF o exención, la que corresponda"),
    "contabilizacion": "el skill compara importe y fecha del documento con los de libros",
    "auditor": "queda al auditor: no se infiere del documento",
}


def _cabecera(ws, fila: int, textos: list[str]) -> None:
    for j, t in enumerate(textos, start=1):
        c = ws.cell(row=fila, column=j, value=t)
        c.font = CABECERA_FONT
        c.fill = CABECERA_FILL
        c.border = BORDE
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _anchos(ws, anchos: dict[int, float]) -> None:
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _hoja_muestra(wb: Workbook, evaluados: list[dict], cols: dict, atributos: list[dict],
                  roles: dict, params: dict, generado: str) -> None:
    """La unica hoja del papel: un elemento por fila y la observacion propuesta.

    **Las columnas de atributos van en blanco a proposito.** El skill no marca
    Ok ni escribe 'auditor' en ellas: las deja para que las rellene quien firma,
    y dice lo que ha visto en la observacion, que es el texto que se copia a
    ForSampling. Decision del auditor el 03/09/2026, y es el modelo de partida.

    Lo que se cuenta y no cabe en la hoja -recuentos, documentos que no son de
    ningun elemento, comparacion con la evaluacion del auditor- se imprime al
    ejecutar. No se pierde: cambia de sitio.
    """
    ws = wb.active
    ws.title = "Análisis muestra"
    if not evaluados:
        ws["A1"] = "La muestra no trae elementos."
        return
    col_pob = list(evaluados[0]["fila"].keys())
    col_pob = [c for c in col_pob if c.lower() not in ("seleccionado",)]
    col_doc = ["Fichero", "Nº factura (doc.)", "Fecha (doc.)", "Base (doc.)", "IVA (doc.)",
               "IRPF (doc.)", "Total (doc.)", "Casa por", "Diferencia importe", "Días doc.–libros"]
    col_att = [f"A{a.get('AtributoId')} {a.get('Nombre')}" for a in atributos]
    cab = col_pob + col_doc + col_att + ["Observación propuesta"]
    ws["A1"] = f"Análisis de la muestra — {params.get('Prueba') or 'prueba de cumplimiento'}"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    def _dia(v):
        d = parse_fecha(v)
        return d.strftime("%d/%m/%Y") if d else (str(v) if v else "—")

    ws["A2"] = (f"Prueba {params.get('MuestraId')} · {params.get('Tipo')} · área {params.get('Area')} · "
                f"ref. {params.get('Referencia')} · ejercicio "
                f"{_dia(params.get('FechaInicioAuditoria'))} — {_dia(params.get('FechaFinAuditoria'))} · "
                f"generado {generado}")
    ws["A2"].font = Font(name=FONT, size=9)
    ws["A3"] = ("Propuesta del asistente: localiza cada documento y comprueba lo que se puede comprobar desde él. "
                "Las columnas de atributos se dejan en blanco para el auditor, que es quien concluye y firma.")
    ws["A3"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws["A4"] = ("Amarillo: la observación señala algo. Este papel no prueba que el documento sea auténtico, ni que "
                "el gasto estuviera autorizado, ni que el registro fuera oportuno: nada de eso está en la factura.")
    ws["A4"].font = Font(name=FONT, italic=True, size=9, color="595959")
    _cabecera(ws, 6, cab)
    fila = 7
    for e in evaluados:
        f, fac, c, res = e["fila"], e["factura"], e["criterios"], e["resultados"]
        valores = []
        for k in col_pob:
            v = f.get(k)
            if k == cols.get("importe"):
                v = parse_importe(v)
            elif k == cols.get("fecha"):
                d = parse_fecha(v)
                v = d.strftime("%d/%m/%Y") if d else v
            valores.append(v)
        if fac:
            casa = ", ".join(x for x, b in (("número", c["numero"]), ("tercero", c.get("tercero")),
                                            (f"importe ({c['importe']})", c["importe"]),
                                            ("fecha", c["fecha"])) if b)
            valores += [fac.get("fichero"), fac.get("numero"), fac.get("fecha"), parse_importe(fac.get("base")),
                        parse_importe(fac.get("iva")), parse_importe(fac.get("irpf")),
                        parse_importe(fac.get("total")), casa, c["diferencia"], c["dias"]]
        else:
            valores += ["— no localizada", None, None, None, None, None, None, "", None, None]
        valores += [None] * len(atributos)          # en blanco: las rellena el auditor
        valores.append(e["observacion"])
        for j, v in enumerate(valores, start=1):
            cell = ws.cell(row=fila, column=j, value=v)
            cell.font = Font(name=FONT, size=9)
            cell.border = BORDE
            cell.alignment = Alignment(vertical="top", wrap_text=j > len(col_pob) + len(col_doc))
            if isinstance(v, float):
                cell.number_format = FORMATO_EURO
        if any(v != "auditor" and not v.startswith("Ok") for v in res.values()):
            ws.cell(row=fila, column=len(cab)).fill = AMARILLO
        fila += 1
    ws.auto_filter.ref = f"A6:{get_column_letter(len(cab))}{fila - 1}"
    ws.freeze_panes = "A7"
    anchos = {i + 1: 14 for i in range(len(cab))}
    for i, k in enumerate(col_pob, start=1):
        if k == cols.get("tercero"):
            anchos[i] = 30
    anchos[len(col_pob) + 1] = 34
    for k in range(len(atributos)):
        anchos[len(col_pob) + len(col_doc) + k + 1] = 12
    anchos[len(cab)] = 95
    _anchos(ws, anchos)


def main() -> int:
    salida_utf8()
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--muestra", required=True)
    p.add_argument("--parametros", required=True)
    p.add_argument("--facturas", required=True)
    p.add_argument("--roles", help="{AtributoId: rol} confirmado por el auditor; sin él, la propuesta automática")
    p.add_argument("--evaluacion", help="exportar_consulta(entidad='evaluacion', id=N), para comparar")
    p.add_argument("--salida", required=True)
    p.add_argument("--generado", required=True, help="fecha de generación, AAAA-MM-DD. Nada lee el reloj")
    args = p.parse_args()

    muestra = cargar_muestra(args.muestra)
    params = cargar_parametros(args.parametros)
    facturas = cargar_facturas(args.facturas)
    atributos = params.get("atributos") or []
    roles = cargar_json(args.roles) if args.roles else proponer_roles(atributos)
    roles = {str(k): v for k, v in roles.items()}
    evaluacion = cargar_evaluacion(args.evaluacion)

    cols, _, _ = columnas_de_muestra(muestra)
    cruce = cruzar(muestra, facturas, cols)
    evaluados = evaluar(cruce, cols, atributos, roles)
    comp = comparar_con_auditor(evaluados, evaluacion, cols, atributos)

    wb = Workbook()
    _hoja_muestra(wb, evaluados, cols, atributos, roles, params, args.generado)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(salida)
    except PermissionError:
        print(f"ERROR: no se puede escribir {salida}. Si está abierto en Excel, ciérralo y repite.")
        return 2

    con_doc = sum(1 for e in evaluados if e["factura"])
    hall = sum(1 for e in evaluados for v in e["resultados"].values() if v != "auditor" and not v.startswith("Ok"))
    print(f"Papel escrito: {salida}")
    print(f"  elementos {len(evaluados)} · con documento {con_doc} · sin documento {len(evaluados) - con_doc}"
          f" · documentos sin elemento {len(cruce['facturas_sin_fila'])} · celdas con hallazgo {hall}")
    for par in cruce.get("candidatos_sueltos") or []:
        print(f"  · POSIBLE DIFERENCIA, no extravío: el elemento {par['id']} quedó sin documento y "
              f"{par['fichero']} sin elemento"
              + (", mismo tercero" if par["tercero"] else "")
              + (f", {par['dias']} día(s) de diferencia" if par["dias"] is not None else "")
              + (f". En libros {par['importe_libros']}, en el documento base {par['base']} / total "
                 f"{par['total']}: diferencia mínima {par['diferencia_minima']}"
                 if par["diferencia_minima"] is not None else "")
              + ". Revísalo a mano: si son el mismo, hay una diferencia real que declarar con "
                "poblacion_id en facturas.json")
    for fac in cruce["facturas_sin_fila"]:
        print(f"  · documento que no es de ningún elemento de la muestra: {fac.get('fichero')}"
              f" ({fac.get('proveedor') or '—'}, {fac.get('total') or '—'})")
    for e in evaluados:
        if any(v != "auditor" and not v.startswith("Ok") for v in e["resultados"].values()):
            ident = e["fila"].get(cols.get("id"), "?")
            print(f"  · elemento {ident}: " + " | ".join(f"A{k}: {v}" for k, v in e["resultados"].items()
                                                         if v != "auditor" and not v.startswith("Ok")))
    if comp:
        r = comp["recuento"]
        print(f"  frente al auditor: coinciden {r['coinciden']} · skill señala/auditor Sí {r['skill_senala_auditor_si']}"
              f" · skill Ok/auditor No {r['skill_ok_auditor_no']} · al auditor {r['auditor']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
