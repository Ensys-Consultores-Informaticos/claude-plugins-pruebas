"""Escribe el papel de trabajo de cancelacion de saldos en Excel.

    python generar_papel.py --entrada extracto.csv \
        --salida "<expediente>/InformesGesia/CancelacionSaldos/Cancelacion Saldos <CLIENTE>.xlsx"

Una hoja "Resumen" con una fila por cuenta, y una hoja por cuenta con el
detalle FECHA / CUENTA / NOMBRE / CONCEPTO / SALDO / INDICE / ORIGEN, en
orden cronologico y con autofiltro. Dos colores, y ninguno mas:

  - gris     (fila entera) el apunte venia ya punteado en la contabilidad
             (columna Indice del .smn): se respeta tal cual
  - amarillo (solo la celda del importe) el apunte NO se ha podido parear:
             es lo que compone el saldo vivo de la cuenta, y es donde el
             auditor tiene que mirar

Lo que este papel cancela no se colorea, venga de pareo directo o de
acumulacion de saldo: esa distincion la da la columna ORIGEN. Antes el
amarillo marcaba la acumulacion y se cambio en la revision del
27/08/2026, porque lo que hay que ver de un vistazo es el pendiente.

No usa formulas: los totales se calculan aqui, en Python, y se escriben
como valor. Este papel tiene que leerse igual en Cowork y en la maquina del
auditor sin depender de si hay LibreOffice o Excel instalado para
recalcular -- a diferencia de un artefacto de Cowork, este es un skill que
tambien corre en local.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from lib_cancelacion import (  # noqa: E402
    MAX_GRUPO_COMBOS,
    MAX_LEFTOVER_FOR_COMBOS,
    ORIGEN_CONTABLE,
    analizar_hallazgos,
    cargar_extracto,
    procesar_extracto,
)

FONT = "Arial"
FORMATO_EURO = '#,##0.00 "€";-#,##0.00 "€";"-"'
FORMATO_FECHA = "dd/mm/yyyy"
AMARILLO = PatternFill("solid", fgColor="FFFF00")
GRIS = PatternFill("solid", fgColor="E7E6E6")
CABECERA_FILL = PatternFill("solid", fgColor="1F4E78")
CABECERA_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
_BORDE_LADO = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_BORDE_LADO, right=_BORDE_LADO, top=_BORDE_LADO, bottom=_BORDE_LADO)

MAX_LISTA = 30  # cuentas detalladas en consola; por encima, se resume


def _hoja_cuenta(wb: Workbook, cuenta: str, res, info: dict) -> dict:
    ws = wb.create_sheet(str(cuenta)[:31])
    nombre_cliente = res["NOMBRE"].iloc[0] if len(res) else ""

    ws["A1"] = "Cancelacion de saldos — Cuenta " + str(cuenta) + " — " + nombre_cliente
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws["A2"] = ("Emparejamiento de facturas y pagos — gris: punteado ya en la "
                "contabilidad; amarillo: importe sin parear, que compone el "
                "saldo vivo de la cuenta")
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="595959")
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")

    cab = ["FECHA", "CUENTA", "NOMBRE", "CONCEPTO", "SALDO", "INDICE", "ORIGEN"]
    fila_cab = 4
    for j, h in enumerate(cab, start=1):
        c = ws.cell(fila_cab, j, h)
        c.font = CABECERA_FONT
        c.fill = CABECERA_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDE

    # Orden cronologico: el mayor se lee en su orden natural (decidido en
    # la revision del 27/08/2026). Los grupos se siguen por la
    # columna INDICE, y el autofiltro permite reordenar al revisar.
    ordenado = res.sort_values("FECHA", kind="stable").reset_index(drop=True)
    fila = fila_cab + 1
    for _, r in ordenado.iterrows():
        valores = [r["FECHA"].to_pydatetime(), r["CUENTA"], r["NOMBRE"], r["CONCEPTO"],
                   float(r["SALDO"]), int(r["INDICE"]), r["ORIGEN"]]
        # Gris a la fila entera del punteo contable; amarillo SOLO a la celda
        # del importe que queda sin parear, que es lo que compone el saldo
        # vivo de la cuenta (revision del 27/08/2026). Lo cancelado
        # por este papel no se colorea, venga de pareo directo o de
        # acumulacion: esa distincion la da la columna ORIGEN, no el color.
        relleno_fila = GRIS if r["ORIGEN"] == ORIGEN_CONTABLE else None
        sin_parear = int(r["INDICE"]) == 0
        for j, v in enumerate(valores, start=1):
            c = ws.cell(fila, j, v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDE
            if j == 1:
                c.number_format = FORMATO_FECHA
            elif j == 5:
                c.number_format = FORMATO_EURO
            elif j in (6, 7):
                c.alignment = Alignment(horizontal="center")
            if j == 5 and sin_parear:
                c.fill = AMARILLO
            elif relleno_fila:
                c.fill = relleno_fila
        fila += 1

    fila_total = fila
    ws.cell(fila_total, 4, "TOTAL").font = Font(name=FONT, bold=True, size=10)
    total = round(float(ordenado["SALDO"].sum()), 2)
    c = ws.cell(fila_total, 5, total)
    c.font = Font(name=FONT, bold=True, size=10)
    c.number_format = FORMATO_EURO
    for j in range(1, 8):
        ws.cell(fila_total, j).fill = TOTAL_FILL

    fila_pend = fila_total + 1
    ws.cell(fila_pend, 4, "Pendiente sin cancelar (INDICE = 0)").font = (
        Font(name=FONT, size=9, italic=True))
    pendiente = round(float(info["suma_indice_0"]), 2)
    c = ws.cell(fila_pend, 5, pendiente)
    c.font = Font(name=FONT, bold=True, size=10)
    c.number_format = FORMATO_EURO

    fila_sig = fila_pend + 1
    descuadre = round(float(info["descuadre_punteo_previo"]), 2)
    if abs(descuadre) > 0.005:
        ws.cell(fila_sig, 4, "Descuadre del punteo contable (grupos previos "
                             "que no suman 0)").font = Font(name=FONT, size=9, italic=True)
        c = ws.cell(fila_sig, 5, descuadre)
        c.font = Font(name=FONT, bold=True, size=10, color="C00000")
        c.number_format = FORMATO_EURO
        fila_sig += 1

    ok = info["coincide_total_con_no_cancelado"] and not info["grupos_con_error"]
    ws.cell(fila_sig, 4, "Verificacion").font = Font(name=FONT, size=9, italic=True)
    c = ws.cell(fila_sig, 5, "OK" if ok else "REVISAR")
    c.font = Font(name=FONT, bold=True, size=10, color="000000" if ok else "C00000")

    anchos = {1: 12, 2: 13, 3: 20, 4: 34, 5: 15, 6: 9, 7: 11}
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A" + str(fila_cab + 1)
    # solo las filas de datos: las de totales quedan fuera del filtro
    ws.auto_filter.ref = "A" + str(fila_cab) + ":G" + str(fila - 1)

    return {
        "cuenta": cuenta,
        "nombre": nombre_cliente,
        "apuntes": len(ordenado),
        "grupos_previos": info["num_grupos_previos"],
        "grupos_nuevos": info["num_grupos_nuevos"],
        "sin_cancelar": info["num_registros_sin_cancelar"],
        "total": total,
        "pendiente": pendiente,
        "descuadre_previo": descuadre,
        "ok": ok,
    }


def _hoja_criterios(ws, h: dict) -> None:
    """Bajo que criterios se ha hecho el papel, y que ha salido.

    Va la primera a proposito: quien abra el libro tiene que saber que esta
    leyendo antes de leerlo. Y describe, no dictamina -- los numeros dicen
    donde mirar; concluir es del auditor.
    """
    def titulo(fila, texto):
        c = ws.cell(fila, 1, texto)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = CABECERA_FILL
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        return fila + 1

    def linea(fila, etiqueta, valor=None, nota=None):
        c = ws.cell(fila, 1, etiqueta)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if valor is not None:
            v = ws.cell(fila, 2, valor)
            v.font = Font(name=FONT, size=10, bold=True)
            v.alignment = Alignment(horizontal="right", vertical="top")
            if isinstance(valor, float):
                v.number_format = FORMATO_EURO
        if nota is not None:
            n = ws.cell(fila, 3, nota)
            n.font = Font(name=FONT, size=9, italic=True, color="595959")
            n.alignment = Alignment(vertical="top", wrap_text=True)
        return fila + 1

    ws["A1"] = "Cancelacion de saldos — criterios y hallazgos"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws.merge_cells("A1:C1")
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 68

    f = 3
    f = titulo(f, "COMO SE HAN FORMADO LOS GRUPOS")
    f = linea(f, "1 · Punteo previo de la contabilidad", None,
              "Si el diario trae la columna Indice, esos grupos se respetan tal cual "
              "y no se rehacen. Un grupo previo que no suma cero se avisa, no se corrige.")
    f = linea(f, "2 · La apertura primero", None,
              "Se acumulan los apuntes de signo contrario hasta dar su importe exacto. "
              "Si ningun conjunto lo da, la apertura queda pendiente: no se fuerza.")
    f = linea(f, "3 · Pareo directo por importe", None,
              "Apuntes de igual importe y signo contrario, uno a uno.")
    f = linea(f, "4 · Acumulacion cronologica y combinaciones", None,
              "Tramos continuos que sumen cero y, sobre lo que quede, subconjuntos de "
              "hasta " + str(MAX_GRUPO_COMBOS) + " apuntes entre un maximo de "
              + str(MAX_LEFTOVER_FOR_COMBOS) + " sin cancelar.")
    f += 1
    f = linea(f, "Fecha que se usa para emparejar", "contable",
              "La del asiento. No se usa el texto del concepto para decidir que apuntes "
              "van juntos.")
    f = linea(f, "Fecha que se usa para informar", "del documento",
              "La que la propia factura lleva escrita en el concepto, cuando la trae. "
              "Solo para los recuentos de abajo.")
    f += 1
    f = linea(f, "LO QUE ESTE PAPEL NO PRUEBA", None,
              "Un grupo es una cancelacion ARITMETICA: sus apuntes suman cero. No es "
              "evidencia documental de que ESE pago liquide ESA factura. Cuando varios "
              "apuntes comparten importe, el emparejamiento concreto es una propuesta.")

    f += 1
    f = titulo(f, "ALCANCE")
    f = linea(f, "Cuentas procesadas", h["cuentas"])
    f = linea(f, "Apuntes", h["apuntes"])
    f = linea(f, "Grupos evaluados", h["grupos_evaluados"])
    if h["grupos_no_evaluables"]:
        f = linea(f, "Grupos no evaluables", h["grupos_no_evaluables"],
                  "No se puede distinguir que lado son documentos y cual pagos: no "
                  "entran en los recuentos de fechas.")

    f += 1
    f = titulo(f, "SALDO DE APERTURA")
    f = linea(f, "Cuentas con apertura detectada", h["aperturas_detectadas"],
              "Se reconoce por estructura: apunte del 1 de enero y el mas antiguo de "
              "la cuenta.")
    f = linea(f, "  cancelada", h["aperturas_canceladas"])
    f = linea(f, "  sigue viva", h["aperturas_vivas"])
    f = linea(f, "  importe vivo", float(h["aperturas_importe_vivo"]))
    f = linea(f, "Cuentas sin apertura detectada", h["cuentas_sin_apertura"],
              "Puede ser una cuenta abierta en el ejercicio, o que el diario no traiga "
              "el asiento de apertura. Si el saldo inicial deberia estar y no aparece, "
              "el extracto no cubre el ejercicio completo.")

    f += 1
    f = titulo(f, "PAGOS ANTERIORES A SU FACTURA")
    if not h["grupos_evaluados"]:
        # Un cero aqui se leeria como "no hay hallazgos" cuando lo que pasa es
        # que no se ha mirado. Se dice, y no se enseña ninguna cifra.
        f = linea(f, "NO SE HA PODIDO EVALUAR", "—",
                  "Para comparar fechas hay que saber que apuntes son documentos y "
                  "cuales pagos, y en estas cuentas no se puede deducir: no hay "
                  "apertura y el saldo es cero. Este apartado NO dice que no haya "
                  "hallazgos: dice que no se han buscado.")
        return
    f = linea(f, "Con la fecha del documento", h["anomalos"],
              "Los que hay que mirar: el pago es anterior a la fecha que la factura "
              "lleva escrita.")
    f = linea(f, "  la pareja venia forzada por el importe", h["anom_forzados"],
              "Ese importe aparece una sola vez a cada lado: no habia emparejamiento "
              "alternativo, asi que no es una eleccion del papel.")
    f = linea(f, "  habia emparejamiento alternativo", h["anom_con_alternativa"],
              "Aqui si se ha elegido entre varios candidatos del mismo importe. Son "
              "los primeros a revisar.")
    f = linea(f, "  grupos de mas de dos apuntes", h["anom_grupos_grandes"])
    f = linea(f, "Lo parecen solo por la fecha de registro", h["solo_fecha_registro"],
              "El pago es anterior al ASIENTO de la factura pero no a su fecha de "
              "documento: es la contabilidad registrando a fin de mes, no una anomalia. "
              "NO se cuentan arriba.")
    if h["facturas_totales"]:
        pct = 100.0 * h["facturas_con_fecha_doc"] / h["facturas_totales"]
        f = linea(f, "Facturas con fecha de documento en el concepto",
                  format(pct, ".0f") + " %",
                  "De esto depende la fiabilidad de los dos recuentos anteriores. Si es "
                  "bajo, la mayoria se comparan contra la fecha del asiento.")

    if h["plazo_mediana"] is not None:
        f += 1
        f = titulo(f, "PLAZO DE PAGO MEDIDO")
        f = linea(f, "Mediana", str(h["plazo_mediana"]) + " dias",
                  "Medido sobre " + str(h["plazo_muestra"]) + " grupos, de la fecha del "
                  "documento a la del pago. Es un hallazgo, no un criterio: no se ha "
                  "usado para emparejar nada.")
        f = linea(f, "Cuartiles", str(h["plazo_p25"]) + " a " + str(h["plazo_p75"]) + " dias")


def _rellenar_resumen(ws, filas: list[dict]) -> None:
    ws["A1"] = "Cancelacion de saldos — resumen por cuenta"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws.merge_cells("A1:J1")

    cab = ["CUENTA", "NOMBRE", "APUNTES", "GRUPOS PREVIOS (CONTABILIDAD)",
           "GRUPOS NUEVOS (ESTE PAPEL)", "SIN CANCELAR", "TOTAL CUENTA",
           "PENDIENTE (INDICE 0)", "DESCUADRE PUNTEO PREVIO", "VERIFICACION"]
    fila_cab = 3
    for j, h in enumerate(cab, start=1):
        c = ws.cell(fila_cab, j, h)
        c.font = CABECERA_FONT
        c.fill = CABECERA_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDE

    fila = fila_cab + 1
    for f in filas:
        valores = [f["cuenta"], f["nombre"], f["apuntes"], f["grupos_previos"],
                   f["grupos_nuevos"], f["sin_cancelar"], f["total"], f["pendiente"],
                   f["descuadre_previo"], "OK" if f["ok"] else "REVISAR"]
        for j, v in enumerate(valores, start=1):
            c = ws.cell(fila, j, v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDE
            if j in (7, 8, 9):
                c.number_format = FORMATO_EURO
            if j == 9 and abs(f["descuadre_previo"]) > 0.005:
                c.font = Font(name=FONT, size=10, bold=True, color="C00000")
            if j == 10 and not f["ok"]:
                c.font = Font(name=FONT, size=10, bold=True, color="C00000")
        fila += 1

    anchos = {1: 13, 2: 22, 3: 10, 4: 15, 5: 15, 6: 12, 7: 16, 8: 18, 9: 15, 10: 13}
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A" + str(fila_cab + 1)
    ws.auto_filter.ref = "A" + str(fila_cab) + ":J" + str(fila - 1)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--entrada", required=True)
    p.add_argument("--salida", required=True)
    args = p.parse_args()

    df = cargar_extracto(args.entrada)
    por_cuenta = procesar_extracto(df)

    wb = Workbook()
    wb.remove(wb.active)
    # las dos primeras hojas se crean ya para fijar el orden y se rellenan al final
    ws_criterios = wb.create_sheet("Criterios y hallazgos")
    ws_resumen = wb.create_sheet("Resumen")

    filas_resumen = []
    for cuenta, (res, info) in por_cuenta.items():
        filas_resumen.append(_hoja_cuenta(wb, cuenta, res, info))
    _rellenar_resumen(ws_resumen, filas_resumen)
    hallazgos = analizar_hallazgos(df, por_cuenta)
    _hoja_criterios(ws_criterios, hallazgos)

    salida = Path(args.salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(salida)
    except PermissionError:
        # Pasa cuando se regenera el papel con el fichero abierto en Excel.
        print("ERROR: no se puede escribir " + str(salida))
        print("       El fichero esta abierto en Excel. Cierralo y repite.")
        return 2

    print("Escrito: " + str(salida))
    print("  " + str(len(filas_resumen)) + " cuenta(s)")
    for f in filas_resumen[:MAX_LISTA]:
        print("  " + str(f["cuenta"]) + " " + str(f["nombre"])[:30].ljust(30)
              + " " + str(f["apuntes"]).rjust(5) + " apuntes | "
              + str(f["grupos_previos"]).rjust(4) + " previos | "
              + str(f["grupos_nuevos"]).rjust(4) + " nuevos | "
              + str(f["sin_cancelar"]).rjust(4) + " sin cancelar | pendiente "
              + format(f["pendiente"], ",.2f") + " €")
    if len(filas_resumen) > MAX_LISTA:
        print("  ... y " + str(len(filas_resumen) - MAX_LISTA)
              + " cuenta(s) mas: ver hoja Resumen")

    con_descuadre = [f for f in filas_resumen if abs(f["descuadre_previo"]) > 0.005]
    if con_descuadre:
        print("  AVISO: " + str(len(con_descuadre)) + " cuenta(s) con punteo "
              "contable que no suma 0 ("
              + ", ".join(str(f["cuenta"]) for f in con_descuadre[:5])
              + ("..." if len(con_descuadre) > 5 else "")
              + "). Se respeta tal cual: lo juzga el auditor.")
    sin_combinatoria = [f for f in filas_resumen
                        if f["sin_cancelar"] > MAX_LEFTOVER_FOR_COMBOS]
    if sin_combinatoria:
        print("  AVISO: en " + str(len(sin_combinatoria)) + " cuenta(s) quedaron "
              "mas de " + str(MAX_LEFTOVER_FOR_COMBOS) + " apuntes sin cancelar ("
              + ", ".join(str(f["cuenta"]) for f in sin_combinatoria[:5])
              + ("..." if len(sin_combinatoria) > 5 else "")
              + "): la busqueda por combinaciones no se intento ahi. Puede ser "
              "una cuenta que no parea por importes (ventas/cobros agregados "
              "por dia, remesas) o que falten periodos en el extracto.")
    if hallazgos["anomalos"]:
        print("  AVISO: " + str(hallazgos["anomalos"]) + " grupo(s) con el pago anterior "
              "a la fecha de la factura (" + str(hallazgos["anom_con_alternativa"])
              + " donde habia emparejamiento alternativo). Otros "
              + str(hallazgos["solo_fecha_registro"]) + " lo parecen solo por la fecha "
              "de registro y no se cuentan. Detalle en la hoja Criterios y hallazgos.")
    if hallazgos["aperturas_vivas"]:
        print("  AVISO: " + str(hallazgos["aperturas_vivas"]) + " apertura(s) sin cancelar, "
              + format(hallazgos["aperturas_importe_vivo"], ",.2f") + " €")
    con_error = [f for f in filas_resumen if not f["ok"]]
    if con_error:
        print("  AVISO: " + str(len(con_error)) + " cuenta(s) no verifican -- revisar "
              "hoja Resumen antes de entregar")
    return 0


if __name__ == "__main__":
    sys.exit(main())
